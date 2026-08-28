import openpyxl
from io import BytesIO

from django.contrib.auth.models import User
from django.core.management import call_command
from django.test import TestCase

from products.models import Product

from .models import Customer, Invoice, InvoiceItem, StockIn, StockOut, Supplier


class AutoInvoiceNumberingTests(TestCase):

    def setUp(self):
        self.user = User.objects.create_superuser("admin", password="pass12345")
        self.client.login(username="admin", password="pass12345")
        self.supplier = Supplier.objects.create(name="Test Supplier")
        self.customer = Customer.objects.create(name="Test Customer")

    def test_stock_in_auto_numbers_and_disables_field(self):
        resp = self.client.get("/inventory/add/")
        self.assertContains(resp, "PUR0000001")
        self.assertContains(resp, "disabled")

        resp = self.client.post("/inventory/add/", {
            "supplier": self.supplier.id, "item_name": "Steel Pipe",
            "quantity": "5", "unit_cost": "9.00", "date": "2026-08-23", "notes": "",
        })
        self.assertRedirects(resp, "/inventory/")

        stock_in = StockIn.objects.get()
        self.assertEqual(stock_in.invoice_number, "PUR0000001")

        invoice = Invoice.objects.get(invoice_number="PUR0000001")
        self.assertEqual(invoice.invoice_type, Invoice.PURCHASE)
        self.assertEqual(invoice.supplier, self.supplier)

        # Second stock-in should increment.
        resp = self.client.get("/inventory/add/")
        self.assertContains(resp, "PUR0000002")

        self.client.post("/inventory/add/", {
            "supplier": self.supplier.id, "item_name": "Another Item",
            "quantity": "1", "unit_cost": "1.00", "date": "2026-08-23", "notes": "",
        })
        self.assertEqual(
            StockIn.objects.filter(invoice_number="PUR0000002").count(), 1
        )

    def test_stock_out_auto_numbers_independently_from_stock_in(self):
        resp = self.client.get("/inventory/stock-out/add/")
        self.assertContains(resp, "INV0000001")

        resp = self.client.post("/inventory/stock-out/add/", {
            "customer": self.customer.id, "item_name": "Steel Pipe",
            "quantity": "2", "unit_price": "20.00", "date": "2026-08-23", "notes": "",
        })
        self.assertRedirects(resp, "/inventory/stock-out/")

        stock_out = StockOut.objects.get()
        self.assertEqual(stock_out.invoice_number, "INV0000001")

        invoice = Invoice.objects.get(invoice_number="INV0000001")
        self.assertEqual(invoice.invoice_type, Invoice.SALE)
        self.assertEqual(invoice.customer, self.customer)

        # PUR and INV sequences don't interfere with each other.
        self.client.post("/inventory/add/", {
            "supplier": self.supplier.id, "item_name": "Steel Pipe",
            "quantity": "5", "unit_cost": "9.00", "date": "2026-08-23", "notes": "",
        })
        self.assertEqual(
            StockIn.objects.filter(invoice_number="PUR0000001").count(), 1
        )


class InvoiceExportTests(TestCase):

    def setUp(self):
        User.objects.create_superuser("admin", password="pass12345")
        self.client.login(username="admin", password="pass12345")
        self.customer = Customer.objects.create(name="Test Customer")

        self.invoice = Invoice.objects.create(
            invoice_number="INV0000001",
            invoice_type=Invoice.SALE,
            customer=self.customer,
            date="2026-08-23",
            subtotal="100.00",
            grand_total="100.00",
            remaining_amount="100.00",
        )
        InvoiceItem.objects.create(
            invoice=self.invoice, item_name="Steel Pipe",
            quantity=2, unit_price="50.00", total="100.00",
        )

    def test_export_downloads_xlsx_with_invoice_rows(self):
        resp = self.client.get("/inventory/invoices/export/")

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment", resp["Content-Disposition"])

        workbook = openpyxl.load_workbook(BytesIO(resp.content))
        sheet = workbook.active

        self.assertEqual(
            [c.value for c in sheet[1]],
            ["Invoice #", "Type", "Date", "Party", "Subtotal",
             "Discount", "Grand Total", "Paid", "Remaining"],
        )
        self.assertEqual(sheet[2][0].value, "INV0000001")
        self.assertEqual(sheet[2][1].value, "Sale")
        self.assertEqual(sheet[2][3].value, "Test Customer")
        self.assertEqual(sheet[2][6].value, 100.0)

    def test_export_respects_search_filter(self):
        Invoice.objects.create(
            invoice_number="INV0000002", invoice_type=Invoice.SALE,
            customer=self.customer, date="2026-08-23",
        )

        resp = self.client.get("/inventory/invoices/export/?search=INV0000002")
        workbook = openpyxl.load_workbook(BytesIO(resp.content))
        sheet = workbook.active

        rows = [row[0].value for row in sheet.iter_rows(min_row=2)]
        self.assertEqual(rows, ["INV0000002"])


class DataTablesWiringTests(TestCase):
    """Confirm listing pages get the DataTables assets/markup, and
    standalone print/export documents deliberately don't."""

    def setUp(self):
        User.objects.create_superuser("admin", password="pass12345")
        self.client.login(username="admin", password="pass12345")

    def test_listing_pages_load_datatables(self):
        pages = [
            "/", "/inventory/", "/inventory/stock-out/", "/inventory/suppliers/",
            "/inventory/customers/", "/inventory/invoices/", "/expenses/",
            "/accounts/users/",
        ]
        for url in pages:
            with self.subTest(url=url):
                resp = self.client.get(url)
                body = resp.content.decode()
                self.assertEqual(resp.status_code, 200)
                self.assertIn('class="datatable"', body)
                self.assertIn("vendor/datatables/dataTables.min.js", body)
                self.assertIn("vendor/jquery-3.7.1.min.js", body)

    def test_print_pages_do_not_load_datatables(self):
        resp = self.client.get("/expenses/print/")
        self.assertNotIn("jquery", resp.content.decode())


class SeedFakeDataCommandTests(TestCase):
    """Runs the seed/clear management commands against the isolated test
    database only, to prove they work correctly and are fully reversible
    before ever touching the real database."""

    def test_seed_then_clear_round_trips_cleanly(self):
        Supplier.objects.create(name="Real Supplier")
        Customer.objects.create(name="Real Customer")
        StockIn.objects.create(
            invoice_number="PUR0000001",
            supplier=Supplier.objects.get(name="Real Supplier"),
            item_name="Real Item", quantity=1, unit_cost="10.00",
            total_amount="10.00", date="2026-08-23",
        )

        call_command("seed_fake_data", count=200)

        self.assertEqual(StockIn.objects.filter(invoice_number__startswith="TESTPUR").count(), 100)
        self.assertEqual(StockOut.objects.filter(invoice_number__startswith="TESTINV").count(), 100)
        self.assertEqual(Invoice.objects.filter(invoice_number__startswith="TESTPUR").count(), 100)
        self.assertEqual(Invoice.objects.filter(invoice_number__startswith="TESTINV").count(), 100)
        self.assertEqual(Supplier.objects.filter(name__startswith="TEST ").count(), 30)
        self.assertEqual(Customer.objects.filter(name__startswith="TEST ").count(), 30)

        # Real, pre-existing data must be untouched.
        self.assertEqual(StockIn.objects.filter(invoice_number="PUR0000001").count(), 1)
        self.assertTrue(Supplier.objects.filter(name="Real Supplier").exists())

        call_command("clear_fake_data")

        self.assertEqual(StockIn.objects.filter(invoice_number__startswith="TESTPUR").count(), 0)
        self.assertEqual(StockOut.objects.filter(invoice_number__startswith="TESTINV").count(), 0)
        self.assertEqual(Invoice.objects.filter(invoice_number__startswith="TESTPUR").count(), 0)
        self.assertEqual(Invoice.objects.filter(invoice_number__startswith="TESTINV").count(), 0)
        self.assertEqual(Supplier.objects.filter(name__startswith="TEST ").count(), 0)
        self.assertEqual(Customer.objects.filter(name__startswith="TEST ").count(), 0)

        # Real data still untouched after cleanup.
        self.assertEqual(StockIn.objects.filter(invoice_number="PUR0000001").count(), 1)
        self.assertTrue(Supplier.objects.filter(name="Real Supplier").exists())


class ServerSideDataTablesTests(TestCase):
    """Confirms the *_data JSON endpoints actually page on the server -
    the whole point of this change is that the browser never receives
    more than one page's worth of rows at a time."""

    def setUp(self):
        User.objects.create_superuser("admin", password="pass12345")
        self.client.login(username="admin", password="pass12345")

        supplier = Supplier.objects.create(name="Acme Supplier")
        customer = Customer.objects.create(name="Acme Customer")

        for i in range(30):
            StockIn.objects.create(
                invoice_number=f"PUR{i:07d}", supplier=supplier,
                item_name=f"Item {i}", quantity=1, unit_cost="10.00",
                total_amount="10.00", date="2026-08-23",
            )
            StockOut.objects.create(
                invoice_number=f"INV{i:07d}", customer=customer,
                item_name=f"Item {i}", quantity=1, unit_price="20.00",
                total_amount="20.00", date="2026-08-23",
            )
            Invoice.objects.create(
                invoice_number=f"PUR{i:07d}", invoice_type=Invoice.PURCHASE,
                supplier=supplier, date="2026-08-23",
            )
            Invoice.objects.create(
                invoice_number=f"INV{i:07d}", invoice_type=Invoice.SALE,
                customer=customer, date="2026-08-23",
            )

    def test_stock_in_data_pages_on_the_server(self):
        resp = self.client.get(
            "/inventory/data/?draw=1&start=0&length=10&search[value]="
        )
        payload = resp.json()

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(payload["recordsTotal"], 30)
        self.assertEqual(payload["recordsFiltered"], 30)
        self.assertEqual(len(payload["data"]), 10)

    def test_stock_in_data_search_filters_on_the_server(self):
        resp = self.client.get(
            "/inventory/data/?draw=1&start=0&length=10&search[value]=Item 5"
        )
        payload = resp.json()

        self.assertEqual(payload["recordsTotal"], 30)
        self.assertEqual(payload["recordsFiltered"], 1)
        self.assertEqual(len(payload["data"]), 1)

    def test_stock_out_data_pages_on_the_server(self):
        resp = self.client.get(
            "/inventory/stock-out/data/?draw=1&start=20&length=10&search[value]="
        )
        payload = resp.json()

        self.assertEqual(payload["recordsTotal"], 30)
        self.assertEqual(len(payload["data"]), 10)

    def test_invoices_data_includes_both_types(self):
        resp = self.client.get(
            "/inventory/invoices/data/?draw=1&start=0&length=100&search[value]="
        )
        payload = resp.json()

        # 30 purchase invoices (from Stock In) + 30 sale invoices (from Stock Out).
        self.assertEqual(payload["recordsTotal"], 60)


class PosCheckoutTests(TestCase):
    """The POS is a separate way to make a sale, alongside Stock Out -
    it must decrement Product stock, refuse to oversell, always price
    from the server-side Product record (never trust the client), and
    create a normal Sale invoice."""

    def setUp(self):
        User.objects.create_superuser("admin", password="pass12345")
        self.client.login(username="admin", password="pass12345")
        self.customer = Customer.objects.create(name="POS Customer")
        self.product = Product.objects.create(
            name="Steel Pipe", price="100.00", quantity=10,
        )

    def test_checkout_creates_invoice_and_decrements_stock(self):
        resp = self.client.post("/inventory/pos/checkout/", {
            "customer": self.customer.id,
            "product_id": [str(self.product.id)],
            "quantity": ["3"],
            "discount": "0", "paid_amount": "300",
        })

        self.assertRedirects(resp, "/inventory/invoices/")

        self.product.refresh_from_db()
        self.assertEqual(self.product.quantity, 7)

        invoice = Invoice.objects.get(invoice_type=Invoice.SALE, customer=self.customer)
        self.assertEqual(invoice.subtotal, 300)
        self.assertEqual(invoice.grand_total, 300)
        self.assertEqual(invoice.remaining_amount, 0)

        item = InvoiceItem.objects.get(invoice=invoice)
        self.assertEqual(item.item_name, "Steel Pipe")
        self.assertEqual(item.quantity, 3)
        self.assertEqual(item.unit_price, 100)

    def test_checkout_refuses_to_oversell(self):
        resp = self.client.post("/inventory/pos/checkout/", {
            "customer": self.customer.id,
            "product_id": [str(self.product.id)],
            "quantity": ["99"],
            "discount": "0", "paid_amount": "0",
        }, follow=True)

        self.product.refresh_from_db()
        self.assertEqual(self.product.quantity, 10)
        self.assertFalse(Invoice.objects.filter(customer=self.customer).exists())
        self.assertContains(resp, "Only 10")

    def test_checkout_ignores_client_supplied_price(self):
        # The POS form never even sends a price, but confirm the server
        # can't be tricked by one either - it should be silently ignored.
        resp = self.client.post("/inventory/pos/checkout/", {
            "customer": self.customer.id,
            "product_id": [str(self.product.id)],
            "quantity": ["1"],
            "price": ["1"],  # attempted tampering
            "discount": "0", "paid_amount": "0",
        })

        self.assertRedirects(resp, "/inventory/invoices/")
        item = InvoiceItem.objects.get(invoice__customer=self.customer)
        self.assertEqual(item.unit_price, 100)  # real product price, not 1

    def test_checkout_merges_duplicate_product_lines(self):
        resp = self.client.post("/inventory/pos/checkout/", {
            "customer": self.customer.id,
            "product_id": [str(self.product.id), str(self.product.id)],
            "quantity": ["2", "3"],
            "discount": "0", "paid_amount": "0",
        })

        self.assertRedirects(resp, "/inventory/invoices/")
        self.product.refresh_from_db()
        self.assertEqual(self.product.quantity, 5)  # 10 - (2+3)

        item = InvoiceItem.objects.get(invoice__customer=self.customer)
        self.assertEqual(item.quantity, 5)

    def test_pos_page_renders(self):
        resp = self.client.get("/inventory/pos/")

        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Steel Pipe")
        self.assertContains(resp, "POS Customer")

    def test_checkout_requires_customer(self):
        resp = self.client.post("/inventory/pos/checkout/", {
            "product_id": [str(self.product.id)],
            "quantity": ["1"],
        }, follow=True)

        self.assertContains(resp, "Select a customer")
        self.assertFalse(Invoice.objects.filter(invoice_type=Invoice.SALE).exists())


class EditInvoicePermissionTests(TestCase):
    """Only admins (superusers) may edit an invoice - staff can view and
    print, but the Edit page/link is off-limits to them."""

    def setUp(self):
        User.objects.create_superuser("admin", password="pass12345")
        User.objects.create_user("staff", password="pass12345")

        customer = Customer.objects.create(name="Test Customer")
        self.invoice = Invoice.objects.create(
            invoice_number="INV0000001", invoice_type=Invoice.SALE,
            customer=customer, date="2026-08-23",
            subtotal="100.00", grand_total="100.00", remaining_amount="100.00",
        )

    def test_staff_cannot_open_edit_invoice_page(self):
        self.client.login(username="staff", password="pass12345")

        resp = self.client.get(f"/inventory/invoices/{self.invoice.pk}/edit/")

        self.assertRedirects(resp, "/")

    def test_staff_cannot_submit_edit_invoice(self):
        self.client.login(username="staff", password="pass12345")

        resp = self.client.post(f"/inventory/invoices/{self.invoice.pk}/edit/", {
            "invoice_number": "HACKED0001", "party": "1",
            "date": "2026-08-23", "discount": "0", "paid_amount": "0",
        })

        self.assertRedirects(resp, "/")
        self.invoice.refresh_from_db()
        self.assertEqual(self.invoice.invoice_number, "INV0000001")

    def test_admin_can_open_edit_invoice_page(self):
        self.client.login(username="admin", password="pass12345")

        resp = self.client.get(f"/inventory/invoices/{self.invoice.pk}/edit/")

        self.assertEqual(resp.status_code, 200)

    def test_edit_link_hidden_from_staff_in_invoices_list(self):
        self.client.login(username="staff", password="pass12345")

        resp = self.client.get(
            "/inventory/invoices/data/?draw=1&start=0&length=10&search[value]="
        )
        payload = resp.json()

        self.assertNotIn("Edit", payload["data"][0]["actions"])
        self.assertIn("Print", payload["data"][0]["actions"])

    def test_edit_link_shown_to_admin_in_invoices_list(self):
        self.client.login(username="admin", password="pass12345")

        resp = self.client.get(
            "/inventory/invoices/data/?draw=1&start=0&length=10&search[value]="
        )
        payload = resp.json()

        self.assertIn("Edit", payload["data"][0]["actions"])
